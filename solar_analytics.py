import pandas as pd
import requests
import os
import time
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# =============================================================================
# [MANUAL CONFIGURATION AREA] - EDIT THIS SECTION ONLY
# =============================================================================
# 1. ANALYSIS SETTINGS
TARGET_MONTH = "2026-01"   # Target Period (YYYY-MM)
TILT_FACTOR = 1.35         # Winter Correction Factor (Jan)
SYSTEM_LOSS = 0.85         # Efficiency Assumption (0.85 = 15% System Loss)

# 2. PLANT DATABASE INPUT
# Instructions: Add your plants below using the format:
# {"id": "Name", "file": "data.xlsx", "kwp": 1000, "lat": 41.00, "lon": 28.00},
PLANTS = [
    # [PASTE YOUR DATA HERE] 
    
    
]

# =============================================================================
# ⚙️ SOLAR ANALYTICS ENGINE (CORE LOGIC)
# =============================================================================
def main():
    print("--------------------------------------------------")
    print(f"🚀 SOLAR PERFORMANCE ANALYZER v1.0")
    print(f"📅 Target Period: {TARGET_MONTH}")
    print("--------------------------------------------------")

    # [INTERACTIVE GUIDANCE] Check if database is empty
    if not PLANTS:
        print("\n[!] SYSTEM IDLE: NO DATA FOUND.")
        print("--------------------------------------------------")
        print("⚠️  ACTION REQUIRED:")
        print("1. Open this script in your editor.")
        print("2. Scroll to '[MANUAL CONFIGURATION AREA]'.")
        print("3. Enter your Plant ID, Excel Filename, Capacity, and Coordinates.")
        print("4. Save and run the script again.")
        print("--------------------------------------------------")
        return # Stop execution here

    FINAL_SUMMARY = []
    
    for p in PLANTS:
        if not os.path.exists(p['file']):
            print(f"⚠️  File Not Found: {p['file']} (Skipping)")
            continue
        try:
            # 1. Load & Parse Excel
            df = pd.read_excel(p['file'], header=5)
            # Smart Column Detection
            p_col = [c for c in df.columns if any(k in str(c).lower() for k in ['gerçek', 'üretim', 'active', 'production', 'yield'])][0]
            df = df[['Tarih', p_col]].rename(columns={p_col: 'Yield'})
            
            # 2. Data Sanitization (Remove Duplicates)
            df['Tarih'] = pd.to_datetime(df['Tarih'], errors='coerce')
            df = df.dropna(subset=['Tarih']).groupby('Tarih', as_index=False)['Yield'].sum()
            df = df[df['Tarih'].dt.strftime('%Y-%m') == TARGET_MONTH]
            
            # 3. Unit Normalization
            if not df.empty and df['Yield'].mean() < 100: df['Yield'] *= 1000

            # 4. Satellite API Connection
            url = f"https://api.open-meteo.com/v1/forecast?latitude={p['lat']}&longitude={p['lon']}&start_date={TARGET_MONTH}-01&end_date={TARGET_MONTH}-31&daily=shortwave_radiation_sum&timezone=auto"
            for _ in range(3):
                try:
                    res = requests.get(url, timeout=20).json()
                    break
                except: time.sleep(1)
            
            df_w = pd.DataFrame({'Tarih': pd.to_datetime(res["daily"]["time"]), 'Irr': res["daily"]["shortwave_radiation_sum"]})
            df_w['Irr_kWh'] = df_w['Irr'] / 3.6

            # 5. Physics Calculation
            merged = pd.merge(df, df_w, on='Tarih')
            merged['Exp'] = merged['Irr_kWh'] * TILT_FACTOR * p['kwp'] * SYSTEM_LOSS
            
            real = merged['Yield'].sum()
            exp = merged['Exp'].sum()
            pr = (real / exp) * 100 if exp > 0 else 0
            
            # 6. Quality Filter
            if pr <= 100:
                FINAL_SUMMARY.append({
                    'Plant ID': p['id'], 'Capacity (kWp)': p['kwp'],
                    'Actual Prod. (kWh)': round(real, 1),
                    'Expected Prod. (kWh)': round(exp, 1),
                    'PR (%)': round(pr, 2),
                    'Status': "Excellent" if pr >= 85 else "Standard" if pr >= 70 else "Review Needed"
                })
                print(f"✔️  Processed: {p['id']} | PR: {pr:.2f}%")

        except Exception as e: print(f"❌ Error {p['id']}: {e}")

    # =============================================================================
    # 📊 OUTPUT GENERATION
    # =============================================================================
    if FINAL_SUMMARY:
        df_results = pd.DataFrame(FINAL_SUMMARY)
        
        # A. Traffic Light Chart
        print("\n📊 Generating visualization...")
        colors = ['#27ae60' if x >= 85 else '#f1c40f' if x >= 70 else '#e74c3c' for x in df_results['PR (%)']]
        
        plt.figure(figsize=(14, 8))
        bars = plt.bar(df_results['Plant ID'], df_results['PR (%)'], color=colors, edgecolor='#2c3e50', alpha=0.9)
        
        for bar in bars:
            plt.text(bar.get_x() + bar.get_width()/2., bar.get_height() + 1,
                    f'{bar.get_height():.1f}%', ha='center', va='bottom', fontweight='bold', color='#34495e')

        plt.axhline(y=80, color='#2980b9', linestyle='--', linewidth=2, label='KPI Target (80%)')
        plt.title(f'Solar Fleet Performance Audit - {TARGET_MONTH}', fontsize=16, fontweight='bold', color='#2c3e50', pad=20)
        plt.ylabel('Performance Ratio (%)', fontsize=12, fontweight='bold')
        plt.xticks(rotation=45, ha='right')
        plt.grid(axis='y', linestyle='--', alpha=0.5)
        plt.legend(loc='upper right')
        plt.tight_layout()
        plt.savefig('Fleet_Performance_Chart.png', dpi=300)
        plt.close()

        # B. Excel Report (Smart Save)
        base_name = "SOLAR_PERFORMANCE_REPORT"
        filename = f"{base_name}.xlsx"
        counter = 1
        while True:
            try:
                with open(filename, "a"): pass 
                df_results.to_excel(filename, index=False)
                break
            except PermissionError:
                filename = f"{base_name}_v{counter}.xlsx"
                counter += 1
        
        # C. Styling
        wb = load_workbook(filename); ws = wb.active
        header_style = PatternFill("solid", fgColor="002060")
        font_white = Font(color="FFFFFF", bold=True)
        for cell in ws[1]: cell.fill = header_style; cell.font = font_white; cell.alignment = Alignment(horizontal="center")

        green, yellow, red = PatternFill("solid", fgColor="C6EFCE"), PatternFill("solid", fgColor="FFEB9C"), PatternFill("solid", fgColor="FFC7CE")
        for row in ws.iter_rows(min_row=2, max_row=len(df_results)+1):
            pr_val = row[4].value
            fill = green if pr_val >= 85 else yellow if pr_val >= 70 else red
            row[4].fill = fill; row[5].fill = fill
            for cell in row: cell.alignment = Alignment(horizontal="center"); cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F']: ws.column_dimensions[col].width = 25
        wb.save(filename)
        
        print(f"✅ SUCCESS: Report generated as '{filename}'")
        print("📈 CHART: Saved as 'Fleet_Performance_Chart.png'")
        os.startfile(filename)

if __name__ == "__main__":
    main()
